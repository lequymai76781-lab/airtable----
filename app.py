from pathlib import Path
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
import os
import re
import secrets
import shutil
import sqlite3

from flask import (
    Flask,
    request,
    jsonify,
    render_template,
    send_file,
    send_from_directory,
    session,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash


# =========================================================
# 基础配置
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = BASE_DIR / "data"
DEFAULT_DB_PATH = DEFAULT_DATA_DIR / "choices.db"
FALLBACK_DB_PATH = BASE_DIR / "choices.db"


def can_write_directory(path):
    try:
        path.mkdir(exist_ok=True)
        probe = path / f".write_probe_{secrets.token_hex(4)}"
        probe.write_text("", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return True
    except OSError:
        return False


def resolve_db_path():
    configured_path = os.environ.get("DATABASE_PATH")
    if configured_path:
        db_path = Path(configured_path)
        if not db_path.is_absolute():
            db_path = BASE_DIR / db_path
        return db_path.resolve()

    if DEFAULT_DB_PATH.exists() and not can_write_directory(DEFAULT_DATA_DIR):
        if not FALLBACK_DB_PATH.exists():
            try:
                shutil.copy2(DEFAULT_DB_PATH, FALLBACK_DB_PATH)
            except OSError:
                pass
        return FALLBACK_DB_PATH.resolve()

    return DEFAULT_DB_PATH.resolve()


DB_PATH = resolve_db_path()
DATA_DIR = DB_PATH.parent

DEFAULT_ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")
ADMIN_INITIAL_PASSWORD_FILE = DATA_DIR / "admin_initial_password.txt"

# 页面提示“原则上 3 次”，后端真实限制最多 5 次
MAX_SUBMIT_TIMES_PER_CONTACT = 5
MIN_SUBMIT_INTERVAL_SECONDS = 5
CLEAR_DATA_CONFIRM_TEXT = "确认清空"

PHONE_PATTERN = re.compile(r"^\d{11}$")
EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

TUTOR_NAMES = [
    "陈积银", "窦光华", "付晓静", "付志华", "高立", "韩瑞珍", "侯小琴", "胡家浩",
    "黄海", "姜小凌", "李爱群", "李保存", "李菁", "刘娟", "刘晓丽", "吕永峰",
    "彭肇一", "邵娟", "滕姗姗", "万晓红", "汪蓓", "王创业", "王润斌", "王雷",
    "王相飞", "王真真", "王雪莲", "吴丹", "肖宁", "姚洪磊", "游迎亚", "张德胜",
    "张钢花", "周榕", "邹瑶",
]

app = Flask(
    __name__,
    template_folder=str(BASE_DIR)
)

app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("SESSION_COOKIE_SECURE") == "1",
)
app.json.ensure_ascii = False


@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "same-origin")
    return response


# =========================================================
# 数据库与通用工具
# =========================================================

def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_conn():
    DATA_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    DATA_DIR.mkdir(exist_ok=True)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_name TEXT NOT NULL,
            major TEXT NOT NULL,
            contact TEXT NOT NULL UNIQUE,
            first_tutor TEXT NOT NULL,
            second_tutor TEXT NOT NULL,
            third_tutor TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS submission_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_name TEXT NOT NULL,
            major TEXT NOT NULL,
            contact TEXT NOT NULL,
            first_tutor TEXT NOT NULL,
            second_tutor TEXT NOT NULL,
            third_tutor TEXT NOT NULL,
            action TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_submission_logs_contact_created_at
        ON submission_logs (contact, created_at)
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_submission_logs_identity_created_at
        ON submission_logs (student_name, major, created_at)
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_submissions_updated_at
        ON submissions (updated_at, id)
    """)

    admin = cur.execute(
        "SELECT id FROM admins WHERE username = ?",
        (DEFAULT_ADMIN_USERNAME,),
    ).fetchone()

    if not admin:
        admin_password = DEFAULT_ADMIN_PASSWORD

        if not admin_password:
            if ADMIN_INITIAL_PASSWORD_FILE.exists():
                admin_password = ADMIN_INITIAL_PASSWORD_FILE.read_text(
                    encoding="utf-8"
                ).strip()
            else:
                admin_password = secrets.token_urlsafe(12)
                ADMIN_INITIAL_PASSWORD_FILE.write_text(
                    admin_password,
                    encoding="utf-8",
                )

        cur.execute("""
            INSERT INTO admins (
                username,
                password_hash,
                created_at
            )
            VALUES (?, ?, ?)
        """, (
            DEFAULT_ADMIN_USERNAME,
            generate_password_hash(admin_password),
            now_text(),
        ))

    conn.commit()
    conn.close()


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if session.get("admin_logged_in") is not True:
            return jsonify({
                "code": 401,
                "message": "请先登录管理员账号。",
            }), 401

        return func(*args, **kwargs)

    return wrapper


# =========================================================
# 业务逻辑
# =========================================================

def normalize_contact(contact):
    contact = str(contact or "").strip()
    if EMAIL_PATTERN.fullmatch(contact):
        return contact.lower()
    return contact


def is_valid_contact(contact):
    """
    联系方式限制：
    1. 11 位纯数字手机号
    2. 合法邮箱：xxx@xxx.xxx
    """
    return (
        PHONE_PATTERN.fullmatch(contact) is not None
        or EMAIL_PATTERN.fullmatch(contact) is not None
    )


def parse_time(text):
    try:
        return datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return None


def validate_submission(data):
    student_name = str(data.get("studentName", "")).strip()
    major = str(data.get("major", "")).strip()
    contact = normalize_contact(data.get("contact", ""))
    first_choice = str(data.get("firstChoice", "")).strip()
    second_choice = str(data.get("secondChoice", "")).strip()
    third_choice = str(data.get("thirdChoice", "")).strip()

    if not student_name:
        return None, "请填写姓名。"

    if not major:
        return None, "请填写专业。"

    if not contact:
        return None, "请填写联系方式。"

    if len(student_name) > 50:
        return None, "姓名长度过长，请核对后重新填写。"

    if len(major) > 100:
        return None, "专业名称长度过长，请核对后重新填写。"

    if len(contact) > 120:
        return None, "联系方式长度过长，请核对后重新填写。"

    if not is_valid_contact(contact):
        return None, "联系方式格式不正确，请填写 11 位手机号或有效邮箱。"

    if not first_choice or not second_choice or not third_choice:
        return None, "请完整选择第一、第二、第三志愿导师。"

    if first_choice not in TUTOR_NAMES:
        return None, "第一志愿导师不在导师名单中。"

    if second_choice not in TUTOR_NAMES:
        return None, "第二志愿导师不在导师名单中。"

    if third_choice not in TUTOR_NAMES:
        return None, "第三志愿导师不在导师名单中。"

    if len({first_choice, second_choice, third_choice}) != 3:
        return None, "第一志愿、第二志愿、第三志愿不能选择同一位导师。"

    return {
        "student_name": student_name,
        "major": major,
        "contact": contact,
        "first_tutor": first_choice,
        "second_tutor": second_choice,
        "third_tutor": third_choice,
    }, None


def get_all_submissions():
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            id,
            student_name,
            major,
            contact,
            first_tutor,
            second_tutor,
            third_tutor,
            created_at,
            updated_at
        FROM submissions
        ORDER BY updated_at DESC, id DESC
    """).fetchall()
    conn.close()
    return rows


def compute_tutor_stats():
    rows = get_all_submissions()

    stats_map = {
        tutor: {
            "tutor": tutor,
            "first": 0,
            "second": 0,
            "third": 0,
            "total": 0,
        }
        for tutor in TUTOR_NAMES
    }

    for row in rows:
        first_tutor = row["first_tutor"]
        second_tutor = row["second_tutor"]
        third_tutor = row["third_tutor"]

        if first_tutor in stats_map:
            stats_map[first_tutor]["first"] += 1

        if second_tutor in stats_map:
            stats_map[second_tutor]["second"] += 1

        if third_tutor in stats_map:
            stats_map[third_tutor]["third"] += 1

    stats = []

    for item in stats_map.values():
        item["total"] = item["first"] + item["second"] + item["third"]
        stats.append(item)

    # 排序规则：第一志愿 > 第二志愿 > 第三志愿 > 姓名
    stats.sort(
        key=lambda item: (
            -item["first"],
            -item["second"],
            -item["third"],
            item["tutor"],
        )
    )

    return stats


# =========================================================
# 页面路由
# =========================================================

@app.route("/")
def index():
    return render_template("app.html")


@app.route("/admin")
def admin_page():
    return render_template("admin.html")


@app.route("/武体校徽.jpeg")
def logo():
    return send_from_directory(BASE_DIR, "武体校徽.jpeg")


# =========================================================
# 学生端接口
# =========================================================

@app.route("/api/tutors", methods=["GET"])
def api_tutors():
    return jsonify({
        "code": 200,
        "data": TUTOR_NAMES,
    })


@app.route("/api/public/tutor-stats", methods=["GET"])
def api_public_tutor_stats():
    return jsonify({
        "code": 200,
        "data": compute_tutor_stats(),
    })


@app.route("/api/submit-choice", methods=["POST"])
def api_submit_choice():
    data = request.get_json(silent=True) or {}
    valid_data, error = validate_submission(data)

    if error:
        return jsonify({
            "code": 400,
            "message": error,
        }), 400

    conn = get_conn()
    cur = conn.cursor()

    current_time = now_text()

    submit_count = cur.execute("""
        SELECT COUNT(*) AS total
        FROM submission_logs
        WHERE contact = ?
    """, (
        valid_data["contact"],
    )).fetchone()["total"]

    if submit_count >= MAX_SUBMIT_TIMES_PER_CONTACT:
        conn.close()
        return jsonify({
            "code": 429,
            "message": "该联系方式提交次数已达上限，请联系学院老师处理。",
        }), 429

    latest_log = cur.execute("""
        SELECT created_at
        FROM submission_logs
        WHERE contact = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 1
    """, (
        valid_data["contact"],
    )).fetchone()

    latest_time = parse_time(latest_log["created_at"]) if latest_log else None
    if latest_time and datetime.now() - latest_time < timedelta(seconds=MIN_SUBMIT_INTERVAL_SECONDS):
        conn.close()
        return jsonify({
            "code": 429,
            "message": "提交过于频繁，请稍后再试。",
        }), 429

    existing = cur.execute("""
        SELECT id
        FROM submissions
        WHERE contact = ?
           OR (student_name = ? AND major = ?)
        ORDER BY
            CASE WHEN contact = ? THEN 0 ELSE 1 END,
            updated_at DESC,
            id DESC
        LIMIT 1
    """, (
        valid_data["contact"],
        valid_data["student_name"],
        valid_data["major"],
        valid_data["contact"],
    )).fetchone()

    if existing:
        cur.execute("""
            UPDATE submissions
            SET
                student_name = ?,
                major = ?,
                contact = ?,
                first_tutor = ?,
                second_tutor = ?,
                third_tutor = ?,
                updated_at = ?
            WHERE id = ?
        """, (
            valid_data["student_name"],
            valid_data["major"],
            valid_data["contact"],
            valid_data["first_tutor"],
            valid_data["second_tutor"],
            valid_data["third_tutor"],
            current_time,
            existing["id"],
        ))

        action = "update"
        message = "提交成功！请确认信息无误，后续如需修改请联系学院老师。"
    else:
        cur.execute("""
            INSERT INTO submissions (
                student_name,
                major,
                contact,
                first_tutor,
                second_tutor,
                third_tutor,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            valid_data["student_name"],
            valid_data["major"],
            valid_data["contact"],
            valid_data["first_tutor"],
            valid_data["second_tutor"],
            valid_data["third_tutor"],
            current_time,
            current_time,
        ))

        action = "create"
        message = "提交成功！请确认信息无误，后续如需修改请联系学院老师。"

    cur.execute("""
        INSERT INTO submission_logs (
            student_name,
            major,
            contact,
            first_tutor,
            second_tutor,
            third_tutor,
            action,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        valid_data["student_name"],
        valid_data["major"],
        valid_data["contact"],
        valid_data["first_tutor"],
        valid_data["second_tutor"],
        valid_data["third_tutor"],
        action,
        current_time,
    ))

    conn.commit()
    conn.close()

    return jsonify({
        "code": 200,
        "message": message,
        "data": {
            "action": action,
            "usedSubmitTimes": submit_count + 1,
            "maxSubmitTimes": MAX_SUBMIT_TIMES_PER_CONTACT,
        },
    })


# =========================================================
# 管理员接口
# =========================================================

@app.route("/api/admin/login", methods=["POST"])
def api_admin_login():
    data = request.get_json(silent=True) or {}

    username = str(data.get("username", "")).strip()
    password = str(data.get("password", "")).strip()

    if not username or not password:
        return jsonify({
            "code": 400,
            "message": "请输入管理员账号和密码。",
        }), 400

    conn = get_conn()
    admin = conn.execute(
        "SELECT id, username, password_hash FROM admins WHERE username = ?",
        (username,),
    ).fetchone()
    conn.close()

    if not admin or not check_password_hash(admin["password_hash"], password):
        return jsonify({
            "code": 401,
            "message": "管理员账号或密码错误。",
        }), 401

    session["admin_logged_in"] = True
    session["admin_id"] = admin["id"]
    session["admin_username"] = admin["username"]

    return jsonify({
        "code": 200,
        "message": "管理员登录成功。",
        "data": {
            "username": admin["username"],
        },
    })


@app.route("/api/admin/logout", methods=["POST"])
def api_admin_logout():
    session.clear()

    return jsonify({
        "code": 200,
        "message": "已退出管理员登录。",
    })


@app.route("/api/admin/status", methods=["GET"])
def api_admin_status():
    return jsonify({
        "code": 200,
        "data": {
            "loggedIn": session.get("admin_logged_in") is True,
            "username": session.get("admin_username"),
        },
    })


@app.route("/api/admin/submissions", methods=["GET"])
@admin_required
def api_admin_submissions():
    rows = get_all_submissions()

    data = []

    for row in rows:
        data.append({
            "id": row["id"],
            "studentName": row["student_name"],
            "major": row["major"],
            "contact": row["contact"],
            "firstChoice": row["first_tutor"],
            "secondChoice": row["second_tutor"],
            "thirdChoice": row["third_tutor"],
            "createdAt": row["created_at"],
            "updatedAt": row["updated_at"],
        })

    return jsonify({
        "code": 200,
        "data": data,
    })


@app.route("/api/admin/tutor-stats", methods=["GET"])
@admin_required
def api_admin_tutor_stats():
    return jsonify({
        "code": 200,
        "data": compute_tutor_stats(),
    })


@app.route("/api/admin/logs", methods=["GET"])
@admin_required
def api_admin_logs():
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            id,
            student_name,
            major,
            contact,
            first_tutor,
            second_tutor,
            third_tutor,
            action,
            created_at
        FROM submission_logs
        ORDER BY created_at DESC, id DESC
        LIMIT 500
    """).fetchall()
    conn.close()

    data = []

    for row in rows:
        data.append({
            "id": row["id"],
            "studentName": row["student_name"],
            "major": row["major"],
            "contact": row["contact"],
            "firstChoice": row["first_tutor"],
            "secondChoice": row["second_tutor"],
            "thirdChoice": row["third_tutor"],
            "action": row["action"],
            "createdAt": row["created_at"],
        })

    return jsonify({
        "code": 200,
        "data": data,
    })


@app.route("/api/admin/clear-data", methods=["POST"])
@admin_required
def api_admin_clear_data():
    """
    清空调试数据：
    1. 清空学生提交结果
    2. 清空提交日志
    3. 保留管理员账号
    """
    data = request.get_json(silent=True) or {}
    confirm_text = str(data.get("confirmText", "")).strip()
    password = str(data.get("password", "")).strip()

    if confirm_text != CLEAR_DATA_CONFIRM_TEXT:
        return jsonify({
            "code": 400,
            "message": "请先完成清空数据确认。",
        }), 400

    if not password:
        return jsonify({
            "code": 400,
            "message": "请输入管理员密码后再清空数据。",
        }), 400

    conn = get_conn()
    cur = conn.cursor()

    admin = cur.execute(
        "SELECT password_hash FROM admins WHERE id = ?",
        (session.get("admin_id"),),
    ).fetchone()

    if not admin or not check_password_hash(admin["password_hash"], password):
        conn.close()
        return jsonify({
            "code": 401,
            "message": "管理员密码校验失败，未清空数据。",
        }), 401

    cur.execute("DELETE FROM submissions")
    cur.execute("DELETE FROM submission_logs")

    conn.commit()
    conn.close()

    return jsonify({
        "code": 200,
        "message": "学生提交数据和提交日志已清空，管理员账号已保留。",
    })


# =========================================================
# Excel 导出
# =========================================================

@app.route("/api/admin/export", methods=["GET"])
@admin_required
def api_admin_export():
    stats = compute_tutor_stats()
    submissions = get_all_submissions()

    wb = Workbook()

    title = "武汉体育学院新闻传播学院研究生选择导师情况详细表"
    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Sheet 1：导师志愿统计
    ws1 = wb.active
    ws1.title = "导师志愿统计"

    build_title_area(
        ws=ws1,
        title=title,
        subtitle=f"导师志愿统计｜导出时间：{export_time}",
        end_col=6,
    )

    write_row(ws1, 4, [
        "序号",
        "导师姓名",
        "总选择人数",
        "第一志愿人数",
        "第二志愿人数",
        "第三志愿人数",
    ])

    for index, item in enumerate(stats, start=1):
        write_row(ws1, 4 + index, [
            index,
            item["tutor"],
            item["total"],
            item["first"],
            item["second"],
            item["third"],
        ])

    style_excel_sheet(
        ws=ws1,
        header_row=4,
        column_widths={
            "A": 10,
            "B": 22,
            "C": 18,
            "D": 18,
            "E": 18,
            "F": 18,
        },
    )

    # Sheet 2：学生提交结果
    ws2 = wb.create_sheet("学生提交结果")

    build_title_area(
        ws=ws2,
        title=title,
        subtitle=f"学生提交结果｜导出时间：{export_time}",
        end_col=9,
    )

    write_row(ws2, 4, [
        "序号",
        "学生姓名",
        "专业",
        "联系方式",
        "第一志愿导师",
        "第二志愿导师",
        "第三志愿导师",
        "初次提交时间",
        "最后更新时间",
    ])

    for index, row in enumerate(submissions, start=1):
        write_row(ws2, 4 + index, [
            index,
            row["student_name"],
            row["major"],
            row["contact"],
            row["first_tutor"],
            row["second_tutor"],
            row["third_tutor"],
            row["created_at"],
            row["updated_at"],
        ])

    style_excel_sheet(
        ws=ws2,
        header_row=4,
        column_widths={
            "A": 10,
            "B": 18,
            "C": 24,
            "D": 28,
            "E": 22,
            "F": 22,
            "G": 22,
            "H": 24,
            "I": 24,
        },
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"武汉体育学院新闻传播学院研究生选择导师情况详细表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def write_row(ws, row_index, values):
    for col_index, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=col_index, value=value)


def build_title_area(ws, title, subtitle, end_col):
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_col,
    )

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=end_col,
    )

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(
        name="微软雅黑",
        size=18,
        bold=True,
        color="1F2937",
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = subtitle
    subtitle_cell.font = Font(
        name="微软雅黑",
        size=11,
        color="64748B",
    )
    subtitle_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 10


def style_excel_sheet(ws, header_row, column_widths):
    header_fill = PatternFill("solid", fgColor="1B61C9")
    header_font = Font(
        name="微软雅黑",
        bold=True,
        color="FFFFFF",
        size=11,
    )

    body_font = Font(
        name="微软雅黑",
        color="111827",
        size=11,
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_row + 1}"

    max_row = ws.max_row
    max_col = ws.max_column

    for row_index in range(header_row, max_row + 1):
        ws.row_dimensions[row_index].height = 28

    for row in ws.iter_rows(
        min_row=header_row,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

            if cell.row == header_row:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = body_font

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
    auto_adjust_column_width(ws, column_widths)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def auto_adjust_column_width(ws, minimum_widths=None, max_width=42):
    minimum_widths = minimum_widths or {}

    for col_index in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_index)
        max_text_width = 0

        for column_cells in ws.iter_cols(
            min_col=col_index,
            max_col=col_index,
            min_row=1,
            max_row=ws.max_row,
        ):
            for cell in column_cells:
                if cell.value is None:
                    continue

                text = str(cell.value)
                text_width = sum(2 if ord(char) > 127 else 1 for char in text)
                max_text_width = max(max_text_width, text_width)

        min_width = minimum_widths.get(column_letter, 10)
        ws.column_dimensions[column_letter].width = min(
            max(min_width, max_text_width + 4),
            max_width,
        )


# =========================================================
# 启动项目
# =========================================================

init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
